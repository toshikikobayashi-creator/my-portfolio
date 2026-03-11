#!/usr/bin/env python3
"""
sales-outreach tracker.py
アタックリストの読み込み・ステータス更新・送信記録を管理するスクリプト。

使い方:
  python tracker.py read <excel_path>          # リスト読み込み（未着手のみ表示）
  python tracker.py update <excel_path> <row> <status>  # ステータス更新
  python tracker.py report <excel_path>         # 送信レポート出力
"""

import sys
import json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


def read_attack_list(excel_path: str) -> list[dict]:
    """アタックリストを読み込み、未着手の企業をJSON形式で返す"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # ヘッダー行を取得
    headers = [cell.value for cell in ws[1]]

    companies = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        record["_row"] = row_idx  # Excel上の行番号を保持

        # ステータスが未着手 or Noneの場合のみ対象
        status = record.get("ステータス", "未着手")
        if status in ("未着手", None, ""):
            companies.append(record)

    # 優先度でソート（高 > 中 > 低）
    priority_order = {"高": 0, "中": 1, "低": 2}
    companies.sort(key=lambda x: priority_order.get(x.get("優先度", "低"), 99))

    return companies


def update_status(excel_path: str, row_number: int, status: str,
                  service: str = "", message_text: str = ""):
    """指定行のステータスを更新し、送信記録を追記する"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 必要なカラムがなければ追加
    needed_cols = {
        "ステータス": None,
        "送信日": None,
        "送信サービス": None,
        "返信有無": None,
        "送信文面": None,
    }

    for col_idx, header in enumerate(headers, start=1):
        if header in needed_cols:
            needed_cols[header] = col_idx

    # 存在しないカラムを追加
    next_col = len(headers) + 1
    for col_name, col_idx in needed_cols.items():
        if col_idx is None:
            ws.cell(row=1, column=next_col, value=col_name)
            needed_cols[col_name] = next_col
            next_col += 1

    # データ更新
    ws.cell(row=row_number, column=needed_cols["ステータス"], value=status)
    ws.cell(row=row_number, column=needed_cols["送信日"],
            value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    if service:
        ws.cell(row=row_number, column=needed_cols["送信サービス"], value=service)
    ws.cell(row=row_number, column=needed_cols["返信有無"], value="未返信")
    if message_text:
        # 文字数制限（Excelセルは32767文字まで）
        ws.cell(row=row_number, column=needed_cols["送信文面"],
                value=message_text[:3000])

    wb.save(excel_path)
    print(f"✅ Row {row_number} updated: {status}")


def generate_report(excel_path: str):
    """送信レポートを生成"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    stats = {"未着手": 0, "送信済み": 0, "返信あり": 0, "商談化": 0, "見送り": 0}
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        total += 1
        status = record.get("ステータス", "未着手") or "未着手"
        if status in stats:
            stats[status] += 1

    print("=" * 40)
    print("📊 アタックリスト レポート")
    print("=" * 40)
    print(f"総企業数:   {total}")
    for status, count in stats.items():
        pct = (count / total * 100) if total > 0 else 0
        bar = "█" * int(pct / 5)
        print(f"  {status}: {count}社 ({pct:.0f}%) {bar}")
    print("=" * 40)

    # 返信率
    sent = stats["送信済み"] + stats["返信あり"] + stats["商談化"]
    replied = stats["返信あり"] + stats["商談化"]
    if sent > 0:
        print(f"返信率: {replied}/{sent} = {replied/sent*100:.1f}%")
        print(f"商談化率: {stats['商談化']}/{sent} = {stats['商談化']/sent*100:.1f}%")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1]
    excel_path = sys.argv[2]

    if command == "read":
        companies = read_attack_list(excel_path)
        print(json.dumps(companies, ensure_ascii=False, indent=2))

    elif command == "update":
        if len(sys.argv) < 5:
            print("Usage: tracker.py update <excel_path> <row_number> <status>")
            sys.exit(1)
        row = int(sys.argv[3])
        status = sys.argv[4]
        update_status(excel_path, row, status)

    elif command == "report":
        generate_report(excel_path)

    else:
        print(f"Unknown command: {command}")
        print(__doc__)
        sys.exit(1)
